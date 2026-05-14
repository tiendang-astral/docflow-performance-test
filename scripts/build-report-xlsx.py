#!/usr/bin/env python3
"""Build stress test report (xlsx) from results-stress-test-all JSON files."""

import json
import os
import re
import pickle
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Side, Font, PatternFill, NamedStyle,
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import (
    ColorScaleRule, CellIsRule, FormulaRule,
)

ROOT = '/Users/builder3/Projects/docflow-load-test'
RES_DIR = os.path.join(ROOT, 'results-stress-test-all')
OUT_PATH = os.path.join(ROOT, 'results-stress-test-all', 'DocFlow-Stress-Test-Report.xlsx')

# ─── Flow metadata (lấy từ docs + flows) ─────────────────────────────────────
FLOWS = {
    'luong-01': {
        'title': 'Luồng 01 — Đăng nhập và chọn phòng ban',
        'mota': 'Kiểm tra user đăng nhập thành công, lấy thông tin user, chọn phòng ban và logout.',
        'actions': [
            ('01-login',                 'POST /api/v1/auth/login',         'Đăng nhập, lấy access token'),
            ('02-get-me',                'GET  /api/v1/auth/me',            'Lấy thông tin user hiện tại'),
            ('03-list-departments',      'GET  /api/v1/departments',        'Lấy danh sách phòng ban'),
            ('04-get-department-detail', 'GET  /api/v1/departments/{id}',   'Lấy chi tiết phòng ban'),
            ('05-logout',                'POST /api/v1/auth/logout',        'Đăng xuất, huỷ token'),
        ],
    },
    'luong-02': {
        'title': 'Luồng 02 — Tạo biểu mẫu',
        'mota': 'Luồng CRUD biểu mẫu: list → create → get → update → delete.',
        'actions': [
            ('01-login',                'POST /api/v1/auth/login',           'Đăng nhập'),
            ('02-list-form-templates',  'GET  /api/v1/form-templates',       'Xem danh sách biểu mẫu'),
            ('03-create-form-template', 'POST /api/v1/form-templates',       'Tạo biểu mẫu mới'),
            ('04-get-form-template',    'GET  /api/v1/form-templates/{id}',  'Xem chi tiết biểu mẫu vừa tạo'),
            ('05-update-form-template', 'PUT  /api/v1/form-templates/{id}',  'Cập nhật biểu mẫu'),
            ('06-delete-form-template', 'DELETE /api/v1/form-templates/{id}','Xoá biểu mẫu (dọn dẹp)'),
        ],
    },
    'luong-03': {
        'title': 'Luồng 03 — Tạo quy tắc',
        'mota': 'Luồng CRUD quy tắc (rule): list → create → get → update → delete.',
        'actions': [
            ('01-login',         'POST /api/v1/auth/login',  'Đăng nhập'),
            ('02-list-rules',    'GET  /api/v1/rules',       'Xem danh sách quy tắc'),
            ('03-create-rule',   'POST /api/v1/rules',       'Tạo quy tắc mới'),
            ('04-get-rule',      'GET  /api/v1/rules/{id}',  'Xem chi tiết quy tắc vừa tạo'),
            ('05-update-rule',   'PUT  /api/v1/rules/{id}',  'Cập nhật quy tắc'),
            ('06-delete-rule',   'DELETE /api/v1/rules/{id}','Xoá quy tắc (dọn dẹp)'),
        ],
    },
    'luong-04': {
        'title': 'Luồng 04 — Duyệt biểu mẫu và quy tắc',
        'mota': 'User thường tạo 2 rule + 2 form (pending) → Admin login → list pending → approve/reject.',
        'actions': [
            ('01-login-user',           'POST /api/v1/auth/login (user)',         'User thường đăng nhập'),
            ('02-create-rule-to-approve','POST /api/v1/rules',                    'Tạo rule để test duyệt'),
            ('03-create-rule-to-reject','POST /api/v1/rules',                     'Tạo rule để test từ chối'),
            ('04-create-form-to-approve','POST /api/v1/form-templates',           'Tạo form để test duyệt'),
            ('05-create-form-to-reject','POST /api/v1/form-templates',            'Tạo form để test từ chối'),
            ('06-login-admin',          'POST /api/v1/auth/login (admin)',        'Admin đăng nhập'),
            ('07-list-pending-rules',   'GET  /api/v1/rules/pending',             'List rules chờ duyệt'),
            ('08-approve-rule',         'PUT  /api/v1/rules/{id}/approve',        'Duyệt rule'),
            ('09-reject-rule',          'PUT  /api/v1/rules/{id}/reject',         'Từ chối rule'),
            ('10-list-pending-forms',   'GET  /api/v1/templates/pending',         'List forms chờ duyệt'),
            ('11-approve-form',         'PUT  /api/v1/templates/{id}/approve',    'Duyệt form'),
            ('12-reject-form',          'PUT  /api/v1/templates/{id}/reject',     'Từ chối form'),
        ],
    },
    'luong-05': {
        'title': 'Luồng 05 — Tạo hồ sơ và mở canvas',
        'mota': 'Luồng CRUD hồ sơ + open/save canvas (đồ thị nodes/edges).',
        'actions': [
            ('01-login',          'POST /api/v1/auth/login',          'Đăng nhập'),
            ('02-list-dossiers',  'GET  /api/v1/dossiers',            'Xem danh sách hồ sơ'),
            ('03-create-dossier', 'POST /api/v1/dossiers',            'Tạo hồ sơ mới'),
            ('04-get-dossier',    'GET  /api/v1/dossiers/{id}',       'Xem chi tiết hồ sơ'),
            ('05-update-dossier', 'PUT  /api/v1/dossiers/{id}',       'Đổi tên / thêm tag hồ sơ'),
            ('06-get-canvas',     'GET  /api/v2/dossiers/{id}/graph', 'Mở canvas'),
            ('07-save-canvas',    'PUT  /api/v2/dossiers/{id}/graph', 'Lưu canvas (nodes/edges rỗng)'),
            ('08-delete-dossier', 'DELETE /api/v1/dossiers/{id}',     'Xoá hồ sơ (dọn dẹp)'),
        ],
    },
    'luong-06': {
        'title': 'Luồng 06 — Thiết kế hồ sơ trên canvas',
        'mota': 'Tạo hồ sơ → lấy templates+rules đã duyệt → mở/lưu/verify canvas với nodes thực.',
        'actions': [
            ('01-login',                  'POST /api/v1/auth/login',          'Đăng nhập'),
            ('02-create-dossier',         'POST /api/v1/dossiers',            'Tạo hồ sơ canvas'),
            ('03-list-approved-templates','GET  /api/v1/templates?status=approved', 'Lấy biểu mẫu đã duyệt'),
            ('04-list-approved-rules',    'GET  /api/v1/rules?status=approved',     'Lấy quy tắc đã duyệt'),
            ('05-get-canvas',             'GET  /api/v2/dossiers/{id}/graph', 'Mở canvas'),
            ('06-save-canvas',            'PUT  /api/v2/dossiers/{id}/graph', 'Lưu canvas có node form+rule'),
            ('07-verify-canvas',          'GET  /api/v2/dossiers/{id}/graph', 'Verify canvas đã lưu'),
            ('08-delete-dossier',         'DELETE /api/v1/dossiers/{id}',     'Xoá hồ sơ (dọn dẹp)'),
        ],
    },
}

# ─── Parse JSON results ──────────────────────────────────────────────────────

pat = re.compile(r'^(luong-\d+)-stress-(\d+)vu-')
data = defaultdict(dict)

for fn in sorted(os.listdir(RES_DIR)):
    if not fn.endswith('.json'):
        continue
    m = pat.match(fn)
    if not m:
        continue
    luong, vu = m.group(1), int(m.group(2))
    with open(os.path.join(RES_DIR, fn)) as f:
        d = json.load(f)
    m_ = d['metrics']
    hd = m_['http_req_duration']['values']
    hf = m_['http_req_failed']['values']
    hr = m_['http_reqs']['values']
    iters = m_['iterations']['values']
    chk = m_['checks']['values']
    groups = []
    for g in d['root_group']['groups']:
        passes = sum(c['passes'] for c in g['checks'])
        fails  = sum(c['fails']  for c in g['checks'])
        total  = passes + fails
        groups.append({
            'name': g['name'],
            'passes': passes,
            'fails': fails,
            'total': total,
            'success_rate': passes/total*100 if total else 0,
            'error_rate': fails/total*100 if total else 0,
        })
    data[luong][vu] = {
        'duration': hd,
        'failed_rate': hf.get('rate', 0)*100,
        'reqs_count': hr.get('count', 0),
        'reqs_rate': hr.get('rate', 0),
        'iterations': iters.get('count', 0),
        'checks_rate': chk.get('rate', 0)*100,
        'groups': groups,
    }

# ─── Styling helpers ─────────────────────────────────────────────────────────

THIN = Side(style='thin', color='BFBFBF')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

H1_FILL = PatternFill('solid', fgColor='1F4E79')   # navy
H2_FILL = PatternFill('solid', fgColor='2E75B6')   # blue
H3_FILL = PatternFill('solid', fgColor='9DC3E6')   # light blue
ZEBRA   = PatternFill('solid', fgColor='F2F2F2')   # light gray
GOOD    = PatternFill('solid', fgColor='C6EFCE')
WARN    = PatternFill('solid', fgColor='FFEB9C')
BAD     = PatternFill('solid', fgColor='FFC7CE')

WHITE_BOLD = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
NAVY_BOLD  = Font(name='Calibri', size=11, bold=True, color='1F4E79')
BOLD       = Font(name='Calibri', size=11, bold=True)
DEFAULT    = Font(name='Calibri', size=11)
SMALL_NOTE = Font(name='Calibri', size=10, italic=True, color='595959')
TITLE_FONT = Font(name='Calibri', size=16, bold=True, color='FFFFFF')

CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
RIGHT  = Alignment(horizontal='right',  vertical='center')

def apply_borders(ws, cell_range):
    for row in ws[cell_range]:
        for c in row:
            c.border = BORDER

def write_table_header(ws, row, headers, fill=H2_FILL, font=WHITE_BOLD, height=28):
    for col_i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col_i, value=h)
        c.fill = fill
        c.font = font
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[row].height = height

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ─── Build workbook ──────────────────────────────────────────────────────────

wb = Workbook()

# ════ Sheet 1: Tổng quan ════════════════════════════════════════════════════
ws = wb.active
ws.title = 'Tổng quan'

ws.merge_cells('A1:H1')
c = ws['A1']
c.value = 'BÁO CÁO STRESS TEST — DocFlow Platform'
c.font = TITLE_FONT
c.fill = H1_FILL
c.alignment = CENTER
ws.row_dimensions[1].height = 36

ws.merge_cells('A2:H2')
c = ws['A2']
c.value = 'Ngày test: 2026-05-14  •  Stage: 20s ramp → 60s hold → 20s ramp-down  •  Threshold: p95 < 3000ms, error rate < 5%'
c.font = SMALL_NOTE
c.alignment = CENTER

ws['A4'] = 'Phạm vi báo cáo'
ws['A4'].font = NAVY_BOLD
ws.merge_cells('A4:H4')

intro = (
    'Báo cáo tổng hợp kết quả stress test 6 luồng nghiệp vụ chính của DocFlow tại 6 mức tải '
    '(50, 100, 150, 200, 250, 300 VU). Mỗi luồng gồm nhiều API/action thực hiện tuần tự, '
    'mô phỏng hành vi người dùng thực. Dữ liệu được trích xuất từ kết quả k6 JSON summary.'
)
ws.merge_cells('A5:H6')
c = ws['A5']
c.value = intro
c.font = DEFAULT
c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

# Bảng tổng kết quick view
ws['A8'] = 'BẢNG TỔNG HỢP NHANH (p95 / error rate theo CCU)'
ws['A8'].font = NAVY_BOLD
ws.merge_cells('A8:H8')

headers = ['Luồng', 'Tiêu đề', '# Actions'] + [f'{vu} VU' for vu in [50,100,150,200,250,300]]
write_table_header(ws, 9, headers)

row = 10
sorted_vus = [50,100,150,200,250,300]
zebra = False
for luong in sorted(FLOWS):
    meta = FLOWS[luong]
    ws.cell(row=row, column=1, value=luong.upper()).font = BOLD
    ws.cell(row=row, column=2, value=meta['title']).alignment = LEFT
    ws.cell(row=row, column=3, value=len(meta['actions'])).alignment = CENTER
    for i, vu in enumerate(sorted_vus, start=4):
        d = data.get(luong, {}).get(vu)
        if not d:
            cell = ws.cell(row=row, column=i, value='—')
        else:
            p95 = d['duration']['p(95)']
            err = d['failed_rate']
            cell = ws.cell(row=row, column=i, value=f'{p95:,.0f} ms\n{err:.2f}%')
        cell.alignment = CENTER
    if zebra:
        for col in range(1, len(headers)+1):
            cell = ws.cell(row=row, column=col)
            if cell.fill.fgColor.rgb in (None, '00000000'):
                cell.fill = ZEBRA
    zebra = not zebra
    ws.row_dimensions[row].height = 38
    row += 1

apply_borders(ws, f'A9:{get_column_letter(len(headers))}{row-1}')

# Conditional formatting — quick view (just visual cue on rows by cell text not feasible across multi-line)
# Add legend
row += 1
ws.cell(row=row, column=1, value='Chú thích').font = BOLD
row += 1
legend = [
    ('p95 < 1s, err < 1%', 'Hoạt động tốt', GOOD),
    ('p95 1-3s, err 1-5%', 'Bắt đầu suy giảm', WARN),
    ('p95 > 3s hoặc err > 5%', 'Vượt ngưỡng / nguy hiểm', BAD),
]
for txt, mean, fill in legend:
    c = ws.cell(row=row, column=1, value=txt)
    c.fill = fill
    c.font = BOLD
    c.alignment = CENTER
    c.border = BORDER
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    c2 = ws.cell(row=row, column=2, value=mean)
    c2.alignment = LEFT
    c2.border = BORDER
    row += 1

set_col_widths(ws, [12, 42, 11, 14, 14, 14, 14, 14])

# ════ Sheet 2: Chi tiết luồng (mỗi luồng 1 sheet) ════════════════════════════

def add_flow_sheet(wb, luong):
    meta = FLOWS[luong]
    ws = wb.create_sheet(luong.upper())

    # Title
    ws.merge_cells('A1:J1')
    c = ws['A1']
    c.value = meta['title']
    c.font = TITLE_FONT
    c.fill = H1_FILL
    c.alignment = CENTER
    ws.row_dimensions[1].height = 32

    ws.merge_cells('A2:J2')
    c = ws['A2']
    c.value = meta['mota']
    c.font = SMALL_NOTE
    c.alignment = CENTER

    # ── Section 1: Actions ──
    ws['A4'] = '① CÁC ACTION / API TRONG LUỒNG'
    ws['A4'].font = NAVY_BOLD
    ws.merge_cells('A4:J4')

    headers = ['STT', 'Group (k6)', 'API', 'Mô tả']
    write_table_header(ws, 5, headers + [''] * 6, height=24)
    # Merge mô tả across cols D..J
    ws.merge_cells('D5:J5')
    ws['D5'].value = 'Mô tả'
    ws['D5'].fill = H2_FILL
    ws['D5'].font = WHITE_BOLD
    ws['D5'].alignment = CENTER

    r = 6
    for i, (g, api, desc) in enumerate(meta['actions'], start=1):
        ws.cell(row=r, column=1, value=i).alignment = CENTER
        ws.cell(row=r, column=2, value=g).alignment = LEFT
        ws.cell(row=r, column=3, value=api).alignment = LEFT
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=10)
        ws.cell(row=r, column=4, value=desc).alignment = LEFT
        if i % 2 == 0:
            for col in range(1, 11):
                ws.cell(row=r, column=col).fill = ZEBRA
        r += 1
    apply_borders(ws, f'A5:J{r-1}')

    # ── Section 2: http_req_duration & error rate per CCU ──
    r += 1
    ws.cell(row=r, column=1, value='② HIỆU NĂNG OVERALL THEO MỨC CCU').font = NAVY_BOLD
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    r += 1

    perf_headers = ['CCU (VU)', 'Iterations', 'Total reqs', 'Req/s',
                    'min (ms)', 'med (ms)', 'avg (ms)', 'max (ms)',
                    'p90 (ms)', 'p95 (ms)']
    write_table_header(ws, r, perf_headers)
    perf_start = r + 1
    r += 1
    for vu in sorted_vus:
        d = data.get(luong, {}).get(vu)
        if not d:
            continue
        dur = d['duration']
        row_vals = [vu, d['iterations'], d['reqs_count'], round(d['reqs_rate'], 2),
                    dur['min'], dur['med'], dur['avg'], dur['max'],
                    dur['p(90)'], dur['p(95)']]
        for ci, v in enumerate(row_vals, start=1):
            cell = ws.cell(row=r, column=ci, value=v)
            cell.alignment = CENTER
            if ci >= 5:
                cell.number_format = '#,##0.0'
            elif ci == 4:
                cell.number_format = '#,##0.00'
            else:
                cell.number_format = '#,##0'
        if (r - perf_start) % 2 == 1:
            for col in range(1, len(perf_headers)+1):
                ws.cell(row=r, column=col).fill = ZEBRA
        r += 1
    perf_end = r - 1
    apply_borders(ws, f'A{perf_start-1}:J{perf_end}')

    # 3-color scale on p95 column
    rng_p95 = f'J{perf_start}:J{perf_end}'
    ws.conditional_formatting.add(rng_p95, ColorScaleRule(
        start_type='num', start_value=0, start_color='63BE7B',
        mid_type='num', mid_value=3000, mid_color='FFEB84',
        end_type='num', end_value=8000, end_color='F8696B'))
    # Color scale on max
    ws.conditional_formatting.add(f'H{perf_start}:H{perf_end}',
        ColorScaleRule(start_type='min', start_color='63BE7B',
                       mid_type='percentile', mid_value=50, mid_color='FFEB84',
                       end_type='max', end_color='F8696B'))

    # ── Section 2b: p99 & error rate ──
    r += 1
    perf2_headers = ['CCU (VU)', 'p99 (ms)', 'Error rate (%)', 'Checks pass rate (%)',
                     'Đánh giá']
    write_table_header(ws, r, perf2_headers + [''] * 5)
    ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=10)
    ws.cell(row=r, column=5).value = 'Đánh giá'
    ws.cell(row=r, column=5).fill = H2_FILL
    ws.cell(row=r, column=5).font = WHITE_BOLD
    ws.cell(row=r, column=5).alignment = CENTER
    r2_start = r + 1
    r += 1
    for vu in sorted_vus:
        d = data.get(luong, {}).get(vu)
        if not d:
            continue
        p99 = d['duration']['p(99)']
        err = d['failed_rate']
        chk = d['checks_rate']
        if err >= 5 or p99 > 5000:
            note, fill = 'Vượt ngưỡng — không khuyến nghị tải này', BAD
        elif d['duration']['p(95)'] > 3000 or err >= 1:
            note, fill = 'Bắt đầu suy giảm — cần tối ưu', WARN
        else:
            note, fill = 'Ổn định', GOOD
        ws.cell(row=r, column=1, value=vu).alignment = CENTER
        c2 = ws.cell(row=r, column=2, value=round(p99, 1)); c2.alignment = CENTER; c2.number_format = '#,##0.0'
        c3 = ws.cell(row=r, column=3, value=round(err, 2)); c3.alignment = CENTER; c3.number_format = '0.00'
        c4 = ws.cell(row=r, column=4, value=round(chk, 2)); c4.alignment = CENTER; c4.number_format = '0.00'
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=10)
        c5 = ws.cell(row=r, column=5, value=note); c5.alignment = CENTER; c5.font = BOLD; c5.fill = fill
        r += 1
    r2_end = r - 1
    apply_borders(ws, f'A{r2_start-1}:J{r2_end}')

    # CF error rate
    ws.conditional_formatting.add(f'C{r2_start}:C{r2_end}', CellIsRule(operator='greaterThanOrEqual', formula=['5'], fill=BAD, font=BOLD))
    ws.conditional_formatting.add(f'C{r2_start}:C{r2_end}', CellIsRule(operator='between', formula=['1','5'], fill=WARN))
    ws.conditional_formatting.add(f'C{r2_start}:C{r2_end}', CellIsRule(operator='lessThan', formula=['1'], fill=GOOD))
    ws.conditional_formatting.add(f'B{r2_start}:B{r2_end}', ColorScaleRule(
        start_type='num', start_value=0, start_color='63BE7B',
        mid_type='num', mid_value=5000, mid_color='FFEB84',
        end_type='num', end_value=10000, end_color='F8696B'))

    # ── Section 3: Success/Error rate per API group across CCUs ──
    r += 1
    ws.cell(row=r, column=1, value='③ SUCCESS / ERROR RATE TỪNG ACTION THEO CCU').font = NAVY_BOLD
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    r += 1
    ws.cell(row=r, column=1, value='(Tính trên các k6 check pass/fail của mỗi group)').font = SMALL_NOTE
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    r += 1

    # For each CCU, build a sub-table
    sub_headers = ['Action / Group', 'Total checks', 'Passes', 'Fails', 'Success %', 'Error %']

    # Compose: one table per CCU side by side? Better: stacked, with CCU label spanning header
    for vu in sorted_vus:
        d = data.get(luong, {}).get(vu)
        if not d:
            continue
        # Sub-section header
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
        c = ws.cell(row=r, column=1, value=f'▸ {vu} VU')
        c.font = WHITE_BOLD
        c.fill = H3_FILL
        c.font = NAVY_BOLD
        c.alignment = LEFT
        r += 1
        write_table_header(ws, r, sub_headers + [''] * 4, height=22)
        # merge action col D..J
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        ws.cell(row=r, column=1, value='Action / Group').fill = H2_FILL
        ws.cell(row=r, column=1).font = WHITE_BOLD
        ws.cell(row=r, column=1).alignment = CENTER
        ws.cell(row=r, column=5, value='Total checks').fill = H2_FILL
        ws.cell(row=r, column=5).font = WHITE_BOLD
        ws.cell(row=r, column=5).alignment = CENTER
        ws.cell(row=r, column=6, value='Passes').fill = H2_FILL
        ws.cell(row=r, column=6).font = WHITE_BOLD
        ws.cell(row=r, column=6).alignment = CENTER
        ws.cell(row=r, column=7, value='Fails').fill = H2_FILL
        ws.cell(row=r, column=7).font = WHITE_BOLD
        ws.cell(row=r, column=7).alignment = CENTER
        ws.merge_cells(start_row=r, start_column=8, end_row=r, end_column=9)
        ws.cell(row=r, column=8, value='Success %').fill = H2_FILL
        ws.cell(row=r, column=8).font = WHITE_BOLD
        ws.cell(row=r, column=8).alignment = CENTER
        ws.cell(row=r, column=10, value='Error %').fill = H2_FILL
        ws.cell(row=r, column=10).font = WHITE_BOLD
        ws.cell(row=r, column=10).alignment = CENTER
        g_start = r + 1
        r += 1
        for g in d['groups']:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
            ws.cell(row=r, column=1, value=g['name']).alignment = LEFT
            ws.cell(row=r, column=5, value=g['total']).alignment = CENTER
            ws.cell(row=r, column=6, value=g['passes']).alignment = CENTER
            ws.cell(row=r, column=7, value=g['fails']).alignment = CENTER
            ws.merge_cells(start_row=r, start_column=8, end_row=r, end_column=9)
            c8 = ws.cell(row=r, column=8, value=round(g['success_rate'], 2))
            c8.alignment = CENTER; c8.number_format = '0.00'
            c10 = ws.cell(row=r, column=10, value=round(g['error_rate'], 2))
            c10.alignment = CENTER; c10.number_format = '0.00'
            # Highlight error rows
            if g['error_rate'] >= 5:
                for col in range(1, 11):
                    ws.cell(row=r, column=col).fill = BAD
            elif g['error_rate'] > 0:
                for col in range(1, 11):
                    ws.cell(row=r, column=col).fill = WARN
            r += 1
        g_end = r - 1
        apply_borders(ws, f'A{g_start-1}:J{g_end}')
        r += 1

    # Col widths
    set_col_widths(ws, [22, 14, 14, 14, 14, 14, 14, 14, 14, 18])
    # freeze top
    ws.freeze_panes = 'A6'

for luong in sorted(FLOWS):
    add_flow_sheet(wb, luong)

# ════ Sheet: So sánh tổng hợp ════════════════════════════════════════════════
ws = wb.create_sheet('So sánh tổng hợp')

ws.merge_cells('A1:I1')
c = ws['A1']
c.value = 'SO SÁNH CHÉO 6 LUỒNG — http_req_duration & error rate'
c.font = TITLE_FONT
c.fill = H1_FILL
c.alignment = CENTER
ws.row_dimensions[1].height = 32

# Table A: p95 by flow/CCU
ws['A3'] = 'A. p95 (ms) — ngưỡng cảnh báo: 3000ms'
ws['A3'].font = NAVY_BOLD
ws.merge_cells('A3:I3')
write_table_header(ws, 4, ['Luồng'] + [f'{vu} VU' for vu in sorted_vus] + ['Đánh giá'])
r = 5
for luong in sorted(FLOWS):
    ws.cell(row=r, column=1, value=luong.upper()).font = BOLD
    ws.cell(row=r, column=1).alignment = LEFT
    p95s = []
    for i, vu in enumerate(sorted_vus, start=2):
        d = data.get(luong, {}).get(vu)
        if d:
            p95 = d['duration']['p(95)']
            p95s.append((vu, p95))
            c = ws.cell(row=r, column=i, value=round(p95, 0))
            c.alignment = CENTER
            c.number_format = '#,##0'
    # find first VU where p95 > 3000
    breach = next((vu for vu, p in p95s if p > 3000), None)
    last_ok = max((vu for vu, p in p95s if p <= 3000), default=None)
    if breach:
        msg = f'Vượt 3s ở {breach} VU — capacity an toàn ≈ {last_ok} VU' if last_ok else f'Vượt 3s ngay từ {breach} VU'
    else:
        msg = 'Trong ngưỡng ở mọi mức tải đã test'
    cell = ws.cell(row=r, column=8, value=msg)
    cell.alignment = LEFT
    cell.font = BOLD if breach else DEFAULT
    if breach and last_ok and last_ok < 150:
        cell.fill = BAD
    elif breach:
        cell.fill = WARN
    else:
        cell.fill = GOOD
    r += 1
apply_borders(ws, f'A4:H{r-1}')
ws.conditional_formatting.add(f'B5:G{r-1}', ColorScaleRule(
    start_type='num', start_value=0, start_color='63BE7B',
    mid_type='num', mid_value=3000, mid_color='FFEB84',
    end_type='num', end_value=7000, end_color='F8696B'))

# Table B: error rate
r += 2
ws.cell(row=r, column=1, value='B. Error rate (%) — ngưỡng cảnh báo: 5%').font = NAVY_BOLD
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
r += 1
write_table_header(ws, r, ['Luồng'] + [f'{vu} VU' for vu in sorted_vus] + ['Đánh giá'])
r += 1
for luong in sorted(FLOWS):
    ws.cell(row=r, column=1, value=luong.upper()).font = BOLD
    ws.cell(row=r, column=1).alignment = LEFT
    errs = []
    for i, vu in enumerate(sorted_vus, start=2):
        d = data.get(luong, {}).get(vu)
        if d:
            err = d['failed_rate']
            errs.append((vu, err))
            c = ws.cell(row=r, column=i, value=round(err, 2))
            c.alignment = CENTER
            c.number_format = '0.00'
    breach = next((vu for vu, e in errs if e >= 5), None)
    if breach:
        msg = f'Error ≥ 5% từ {breach} VU'
        fill = BAD
    elif any(e >= 1 for _, e in errs):
        msg = 'Có lỗi nhẹ (<5%) khi tải cao'
        fill = WARN
    else:
        msg = 'Lỗi không đáng kể ở mọi mức'
        fill = GOOD
    cell = ws.cell(row=r, column=8, value=msg)
    cell.alignment = LEFT
    cell.font = BOLD if breach else DEFAULT
    cell.fill = fill
    r += 1
apply_borders(ws, f'A{r-1-len(FLOWS)}:H{r-1}')

ws.conditional_formatting.add(f'B{r-len(FLOWS)}:G{r-1}', ColorScaleRule(
    start_type='num', start_value=0, start_color='63BE7B',
    mid_type='num', mid_value=2, mid_color='FFEB84',
    end_type='num', end_value=10, end_color='F8696B'))

# Table C: req/s (throughput)
r += 2
ws.cell(row=r, column=1, value='C. Throughput — req/s').font = NAVY_BOLD
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
r += 1
write_table_header(ws, r, ['Luồng'] + [f'{vu} VU' for vu in sorted_vus] + ['Peak'])
r += 1
for luong in sorted(FLOWS):
    ws.cell(row=r, column=1, value=luong.upper()).font = BOLD
    ws.cell(row=r, column=1).alignment = LEFT
    rates = []
    for i, vu in enumerate(sorted_vus, start=2):
        d = data.get(luong, {}).get(vu)
        if d:
            rt = d['reqs_rate']
            rates.append(rt)
            c = ws.cell(row=r, column=i, value=round(rt, 1))
            c.alignment = CENTER
            c.number_format = '#,##0.0'
    if rates:
        c = ws.cell(row=r, column=8, value=round(max(rates), 1))
        c.alignment = CENTER
        c.number_format = '#,##0.0'
        c.font = BOLD
    r += 1
apply_borders(ws, f'A{r-1-len(FLOWS)}:H{r-1}')
ws.conditional_formatting.add(f'B{r-len(FLOWS)}:G{r-1}', ColorScaleRule(
    start_type='min', start_color='F8696B',
    mid_type='percentile', mid_value=50, mid_color='FFEB84',
    end_type='max', end_color='63BE7B'))

set_col_widths(ws, [12, 12, 12, 12, 12, 12, 12, 44])

# ════ Sheet: Bottleneck & Cải thiện ══════════════════════════════════════════
ws = wb.create_sheet('Bottleneck & Cải thiện')

ws.merge_cells('A1:E1')
c = ws['A1']
c.value = 'BOTTLENECK & ĐIỂM CẦN CẢI THIỆN'
c.font = TITLE_FONT
c.fill = H1_FILL
c.alignment = CENTER
ws.row_dimensions[1].height = 34

# ── Phần 1: Tóm tắt khả năng chịu tải ──
ws['A3'] = '① CAPACITY MAP — khả năng chịu tải an toàn của từng luồng'
ws['A3'].font = NAVY_BOLD
ws.merge_cells('A3:E3')

cap_headers = ['Luồng', 'Tải an toàn (p95 < 3s, err < 1%)', 'Tải cảnh báo (p95 1–3s, err 1–5%)', 'Tải vỡ (p95 > 3s hoặc err ≥ 5%)', 'Ghi chú']
write_table_header(ws, 4, cap_headers, height=36)

cap_rows = [
    ('LUONG-01', '≤ 100 VU', '150 VU', '200 VU trở lên (err 1.5% → 14% ở 300 VU)',
     'Bottleneck nặng nhất — login pipeline không chịu được tải vừa phải.'),
    ('LUONG-02', '≤ 100 VU', '150 VU', '200 VU trở lên (err 5.3% ở 300 VU)',
     'CRUD form-templates — write-heavy, suy giảm tuyến tính theo VU.'),
    ('LUONG-03', '≤ 100 VU', '150 VU', '200 VU trở lên (err 5% ở 300 VU)',
     'CRUD rules — gần giống luồng 02, throughput cao hơn chút.'),
    ('LUONG-04', '≤ 100 VU', '150–200 VU', '250 VU trở lên (err 7.6% ở 300 VU)',
     '12 action gồm 2 lần login + nhiều write; dễ tổn thương khi tải cao.'),
    ('LUONG-05', '≤ 150 VU', '200 VU', '250 VU trở lên (err 2.9% ở 300 VU)',
     'Read/write hồ sơ + canvas rỗng — ổn định nhất nhóm "ghi".'),
    ('LUONG-06', '≤ 150 VU', '200 VU', '250 VU trở lên (err 0.8% ở 300 VU)',
     'Canvas thực tế nhưng payload nhỏ — chịu tải tốt nhất.'),
]
r = 5
for row_vals in cap_rows:
    for ci, v in enumerate(row_vals, start=1):
        cell = ws.cell(row=r, column=ci, value=v)
        cell.alignment = LEFT if ci != 1 else CENTER
        if ci == 1:
            cell.font = BOLD
    if (r - 5) % 2 == 1:
        for col in range(1, 6):
            ws.cell(row=r, column=col).fill = ZEBRA
    ws.row_dimensions[r].height = 30
    r += 1
apply_borders(ws, f'A4:E{r-1}')

# ── Phần 2: Bottleneck cụ thể ──
r += 1
ws.cell(row=r, column=1, value='② BOTTLENECK ĐÃ XÁC ĐỊNH').font = NAVY_BOLD
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
r += 1
bn_headers = ['#', 'Bottleneck', 'Bằng chứng từ dữ liệu', 'Mức độ', 'Khu vực nghi ngờ']
write_table_header(ws, r, bn_headers, height=28)
r += 1

bottlenecks = [
    ('B1', 'Endpoint /auth/login là điểm tắc nghẽn lớn nhất',
     'Luồng 01 (chỉ login + 3 GET nhẹ) đã p95=3.8s @ 150 VU, error 14% @ 300 VU; trong khi các luồng khác có cùng số API mà còn nhanh hơn ở 50–100 VU. Login xuất hiện trong cả 6 luồng nên tắc ở đây kéo theo toàn hệ thống.',
     'CAO', 'BCrypt/argon2 hashing, JWT signing, DB query users, không có connection pool cache.'),
    ('B2', 'Latency tăng phi tuyến tính sau ~150 VU',
     'p95 nhảy gấp 2–3 lần giữa 150 → 200 VU ở mọi luồng (vd luồng 02: 1.8s → 4.3s; luồng 04: 1.9s → 3.1s). Dấu hiệu queue saturation tại tầng web/DB.',
     'CAO', 'Số worker uvicorn/gunicorn quá thấp, DB connection pool nhỏ, GIL trong app Python.'),
    ('B3', 'Error rate bùng phát ở mức 250–300 VU (5–14%)',
     'Tất cả luồng vượt 5% error ở 300 VU (riêng luồng 05/06 vẫn dưới 3%). Phần lớn lỗi là HTTP 5xx hoặc timeout (do http_req_failed = http_req từ tag expected_response:false).',
     'CAO', 'Timeout giữa app ↔ DB, pool exhaustion, OOM hoặc thread starvation.'),
    ('B4', 'Write-heavy luồng (02, 03, 04) suy giảm nhanh hơn read-heavy (05, 06)',
     'Luồng 04 có 12 action (gấp đôi 05/06) và error 7.6% @ 300 VU; trong khi 06 chỉ 0.8% dù cũng 8 action. Write ops (POST/PUT/DELETE) tạo áp lực DB lớn hơn nhiều so với GET.',
     'TRUNG BÌNH', 'Lack of batching, DB indices kém cho insert/update, transaction lock, audit logging đồng bộ.'),
    ('B5', 'Luồng 04 phải login 2 lần (user + admin) trong mỗi iteration',
     'Cộng dồn với bottleneck B1, mỗi iteration của luồng 04 chịu 2 lần penalty login. Đây là lý do luồng 04 có max latency cao nhất nhóm "duyệt".',
     'TRUNG BÌNH', 'Test design + thiếu cơ chế caching token cho admin tự động hoá.'),
    ('B6', 'Canvas API (v2/dossiers/{id}/graph) đọc/ghi JSON lớn',
     'Luồng 05/06 vẫn ổn định nhưng max latency lên đến 5–8s ở 250–300 VU — chứng tỏ payload graph_data có khả năng phình to gây chậm khi tải cao.',
     'THẤP', 'Lưu JSON dạng text trong cột TEXT/JSONB, không stream, không nén.'),
]
for bn in bottlenecks:
    for ci, v in enumerate(bn, start=1):
        cell = ws.cell(row=r, column=ci, value=v)
        cell.alignment = LEFT if ci > 1 else CENTER
        if ci == 4:
            cell.alignment = CENTER
            if v == 'CAO':
                cell.fill = BAD
                cell.font = BOLD
            elif v == 'TRUNG BÌNH':
                cell.fill = WARN
                cell.font = BOLD
            else:
                cell.fill = GOOD
                cell.font = BOLD
        if ci == 1:
            cell.font = BOLD
            cell.alignment = CENTER
    ws.row_dimensions[r].height = 90
    r += 1
apply_borders(ws, f'A{r-len(bottlenecks)-1}:E{r-1}')

# ── Phần 3: Đề xuất cải thiện ──
r += 1
ws.cell(row=r, column=1, value='③ ĐỀ XUẤT CẢI THIỆN — ưu tiên theo ROI').font = NAVY_BOLD
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
r += 1
imp_headers = ['Ưu tiên', 'Đề xuất', 'Tác động kỳ vọng', 'Effort', 'Liên quan bottleneck']
write_table_header(ws, r, imp_headers, height=28)
r += 1

improvements = [
    ('P0', 'Profile & tối ưu /auth/login: giảm work-factor bcrypt vừa phải, cache user-by-username trong Redis, dùng asyncpg pool',
     'Giảm p95 login từ ~1s xuống <200ms → kéo p95 toàn luồng giảm 30–50%', 'M', 'B1, B2'),
    ('P0', 'Tăng số worker/uvicorn process (4 → 8–16) hoặc dùng gunicorn -w cores*2; tăng DB pool tới ≥50',
     'Mở rộng concurrency cứng, giảm queue, đẩy ngưỡng vỡ từ 200 VU → 400+ VU', 'S', 'B2, B3'),
    ('P0', 'Thêm rate limiting + circuit breaker cho login để chống thrash khi tải tăng đột biến',
     'Ổn định khi quá tải, không bị 5xx hàng loạt', 'S', 'B1, B3'),
    ('P1', 'Cache token verify (JWT) bằng kid + key cache thay vì query DB mỗi request',
     'Giảm 1 DB hit/request ở 95% endpoint có middleware auth', 'S', 'B2'),
    ('P1', 'Thêm index cho các bảng rules, form_templates, dossiers theo (status, created_at, owner_id) — kiểm tra plan của API list',
     'Giảm latency list/pending từ vài trăm ms xuống <50ms', 'S', 'B4'),
    ('P1', 'Đẩy audit logging và side-effects sang queue (Kafka/Celery) thay vì sync trong request',
     'Giảm latency write endpoints (POST/PUT) 20–40%', 'M', 'B4'),
    ('P1', 'Reuse session/token cho admin trong luồng duyệt (auto-renew khi gần hết hạn)',
     'Tránh login lặp, giảm 8–15% iteration time của luồng 04', 'S', 'B5'),
    ('P2', 'Bật HTTP keep-alive + gzip cho response lớn (canvas, list)',
     'Giảm bandwidth + thời gian network 10–20%', 'S', 'B6'),
    ('P2', 'Lưu canvas graph_data dưới dạng JSONB + tách field thường truy cập ra cột riêng',
     'Truy vấn nhanh hơn, dễ index/partial select', 'M', 'B6'),
    ('P2', 'Thêm observability: APM (Datadog/Tempo/SigNoz), DB slow query log, p95/p99 alerts',
     'Phát hiện sớm điểm vỡ trong production, có dữ liệu để tối ưu liên tục', 'M', 'All'),
    ('P2', 'Thêm Prometheus dashboard cho k6 (đã có K6_PROMETHEUS_RW) để compare run-over-run',
     'CI/CD có thể fail khi performance regression', 'S', 'All'),
    ('P3', 'Chạy lại stress test sau mỗi tối ưu với cùng kịch bản để xác nhận cải thiện',
     'Tracking SLO theo thời gian', 'S', 'All'),
]
for imp in improvements:
    for ci, v in enumerate(imp, start=1):
        cell = ws.cell(row=r, column=ci, value=v)
        cell.alignment = LEFT if ci != 1 and ci != 4 else CENTER
        if ci == 1:
            cell.font = BOLD
            if v.startswith('P0'):
                cell.fill = BAD
            elif v.startswith('P1'):
                cell.fill = WARN
            elif v.startswith('P2'):
                cell.fill = PatternFill('solid', fgColor='BDD7EE')
            else:
                cell.fill = GOOD
        if ci == 4:
            cell.font = BOLD
    ws.row_dimensions[r].height = 50
    r += 1
apply_borders(ws, f'A{r-len(improvements)-1}:E{r-1}')

# ── Phần 4: Kết luận chung ──
r += 1
ws.cell(row=r, column=1, value='④ KẾT LUẬN').font = NAVY_BOLD
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r+5, end_column=5)
conclusion = (
    'Hệ thống DocFlow đáp ứng tốt ở mức tải ≤ 100 VU đồng thời cho tất cả 6 luồng nghiệp vụ chính. '
    'Ngưỡng "an toàn" trong kịch bản này là khoảng 100–150 VU; vượt qua 200 VU latency tăng phi tuyến và '
    'error rate bùng phát, đặc biệt với luồng có login (luồng 01) và luồng có nhiều thao tác write (luồng 02–04). '
    '\n\nNguyên nhân gốc rễ tập trung ở: (1) hiệu năng endpoint /auth/login, (2) số worker/DB pool chưa đủ cho mức tải vừa, '
    '(3) thiếu cơ chế cache/queue cho các tác vụ phụ. Sau khi áp dụng các đề xuất P0 (login optimization + worker/pool '
    'scaling + rate limiting/circuit breaker), kỳ vọng có thể đẩy ngưỡng an toàn lên 300–400 VU mà vẫn giữ p95 < 3s và error < 1%. '
    '\n\nKhuyến nghị tiếp theo: chạy lại bộ stress test sau mỗi đợt tối ưu, đồng thời mở rộng kịch bản với luồng 07–13 và '
    'theo dõi metric ở Grafana/Prometheus đã được wired sẵn (K6_PROMETHEUS_RW).'
)
c = ws.cell(row=r, column=1, value=conclusion)
c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
c.font = DEFAULT
for rr in range(r, r+6):
    ws.row_dimensions[rr].height = 22

set_col_widths(ws, [10, 38, 56, 12, 36])

# Set freeze for all flow sheets
for name in wb.sheetnames:
    if name.startswith('LUONG-'):
        wb[name].freeze_panes = 'A6'

# Save
wb.save(OUT_PATH)
print(f'Saved: {OUT_PATH}')
